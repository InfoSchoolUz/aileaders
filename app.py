"""
AI Leaders PINFL Checker — Premium Edition
Developer: Azamat Madrimov
requirements.txt:
    streamlit
    openpyxl
    pandas
    requests
"""
import re
import time
import requests
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────
API_URL     = "https://aileaders.uz/api/v1/check/certificates"
DELAY_SEC   = 3.5   # Rate limit: 20 req/min → xavfsiz 3.5s
SKIP_SHEETS = {"ЖАМИ СЕРТИФИКАТ ОЛГАНЛАР", "Лист1"}
REPORT_COLS = [
    "Tekshiruv holati", "Saytdagi F.I.Sh.", "Email",
    "Kurslar soni", "Yakunlangan kurslar", "Kurslar tafsiloti", "Izoh"
]

# ──────────────────────────────────────────
# CURL PARSER
# ──────────────────────────────────────────
def parse_curl(curl_text: str) -> dict:
    result = {"cookie": "", "user_agent": ""}
    for pattern in [r"-b\s+'([^']+)'", r'-b\s+"([^"]+)"',
                    r"-H\s+'cookie:\s*([^']+)'", r'-H\s+"cookie:\s*([^"]+)"']:
        m = re.search(pattern, curl_text, re.IGNORECASE)
        if m:
            result["cookie"] = m.group(1).strip()
            break
    for pattern in [r"-H\s+'user-agent:\s*([^']+)'", r'-H\s+"user-agent:\s*([^"]+)"']:
        m = re.search(pattern, curl_text, re.IGNORECASE)
        if m:
            result["user_agent"] = m.group(1).strip()
            break
    return result

# ──────────────────────────────────────────
# EXCEL O'QISH
# ──────────────────────────────────────────
def read_excel(file) -> pd.DataFrame:
    xls = pd.ExcelFile(file, engine="openpyxl")
    frames = []
    rid = 1
    for sheet in xls.sheet_names:
        if sheet in SKIP_SHEETS:
            continue
        try:
            df = pd.read_excel(xls, sheet_name=sheet, header=1, engine="openpyxl")
        except Exception:
            continue
        df.columns = df.columns.map(str).str.strip()
        pinfl_col = next(
            (c for c in df.columns if "ПИНФЛ" in str(c).upper() or "PINFL" in str(c).upper()), None
        )
        if pinfl_col is None:
            continue
        df["_PINFL_COL_"] = pinfl_col
        df["_RID_"] = range(rid, rid + len(df))
        rid += len(df)
        df["PINFL"] = (
            df[pinfl_col].astype(str).str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )
        df["Maktab"] = sheet
        df = df[df["PINFL"].str.len() >= 10].copy()
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

# ──────────────────────────────────────────
# PINFL TEKSHIRISH — rate limit qayta urinish bilan
# ──────────────────────────────────────────
def check_pinfl(pinfl: str, session: requests.Session) -> dict:
    empty = {"holat": "", "ism": "", "email": "", "kurslar": 0,
             "yakunlangan": 0, "kurs_tafsiloti": "", "xato": ""}
    try:
        r = session.get(API_URL, params={"pinfl": pinfl}, timeout=15)

        if r.status_code == 200:
            try:
                data = r.json()
            except Exception:
                return {**empty, "holat": "⚠️ JSON xato", "xato": r.text[:80]}

            courses   = data.get("courses", [])
            completed = sum(1 for c in courses if c.get("isCompleted"))

            if completed > 0:
                holat = "✅ Sertifikat olgan"
            elif len(courses) > 0:
                holat = "⚠️ Kurs bor, sertifikat olinmagan"
            else:
                holat = "⚠️ PINFL bor, kurs topilmadi"

            kurslar = []
            for c in courses:
                nomi     = (c.get("courseName") or c.get("name") or c.get("title") or "Noma'lum kurs")
                hamkor   = (c.get("partner") or c.get("partnerName") or c.get("provider") or "")
                progress = (c.get("progress") or c.get("progressPercent") or "")
                davom    = (c.get("duration") or c.get("approxTotalCourseHrs") or "")
                yozildi  = (c.get("enrolledAt") or c.get("createdAt") or "")
                tugaldi  = (c.get("completedAt") or "")
                ochirildi= (c.get("deletedAt") or "")
                tugatdimi= "Ha" if c.get("isCompleted") else "Yo'q"
                qator = f"Kurs: {nomi}"
                if hamkor:   qator += f" | Hamkor: {hamkor}"
                if progress != "": qator += f" | Progress: {progress}%"
                if davom:    qator += f" | Davomiylik: {davom}h"
                if yozildi:  qator += f" | Yozilgan: {yozildi}"
                if tugaldi:  qator += f" | Tugallangan: {tugaldi}"
                if ochirildi: qator += f" | O'chirilgan: {ochirildi}"
                qator += f" | Sertifikat: {tugatdimi}"
                kurslar.append(qator)

            return {
                "holat": holat,
                "ism":   data.get("fullName", ""),
                "email": data.get("email", ""),
                "kurslar":     len(courses),
                "yakunlangan": completed,
                "kurs_tafsiloti": "\n".join(kurslar),
                "xato": "",
            }

        elif r.status_code == 404:
            return {**empty, "holat": "❌ Ro'yxatdan o'tmagan"}
        elif r.status_code == 401:
            return {**empty, "holat": "🔐 Cookie eskirgan", "xato": "Cookie yangilang"}
        elif r.status_code == 429:
            return {**empty, "holat": "⏳ Rate limit", "xato": "limit"}
        else:
            return {**empty, "holat": f"🔴 Server xatosi: {r.status_code}", "xato": r.text[:80]}
    except Exception as e:
        return {**empty, "holat": "🔴 Xato", "xato": str(e)[:80]}


def check_pinfl_safe(pinfl: str, session: requests.Session, status_el, i: int, total: int) -> dict:
    """Rate limit bo'lsa 3 marta qayta urinadi."""
    for urinish in range(3):
        res = check_pinfl(pinfl, session)
        if res["holat"] != "⏳ Rate limit":
            return res
        wait = 20 + urinish * 10
        for s in range(wait, 0, -1):
            status_el.markdown(
                f'<div class="status-bar">⏳ {i}/{total} — Rate limit! {s}s kutilmoqda... (urinish {urinish+1}/3)</div>',
                unsafe_allow_html=True
            )
            time.sleep(1)
    return {**{"holat": "⏳ Rate limit (3 marta urinildi)", "ism": "", "email": "",
               "kurslar": 0, "yakunlangan": 0, "kurs_tafsiloti": "", "xato": "limit"}}

# ──────────────────────────────────────────
# EXCEL STYLE
# ──────────────────────────────────────────
def style_excel(writer):
    wb = writer.book
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        hfill  = PatternFill("solid", fgColor="1F4E78")
        hfont  = Font(color="FFFFFF", bold=True)
        thin   = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

def build_report_excel(df_all, result_df, summary) -> BytesIO:
    export_df  = df_all.copy()
    result_map = result_df.set_index("_RID_")[REPORT_COLS].to_dict("index")
    for col in REPORT_COLS:
        export_df[col] = export_df["_RID_"].map(lambda x, c=col: result_map.get(x, {}).get(c, ""))
    for col in ["_RID_", "_PINFL_COL_"]:
        if col in export_df.columns:
            export_df = export_df.drop(columns=[col])
    cols = list(export_df.columns)
    for col in REPORT_COLS:
        if col in cols:
            cols.remove(col)
    pinfl_index = next(
        (i for i, c in enumerate(cols) if "ПИНФЛ" in str(c).upper() or c == "PINFL"), 0
    )
    final_cols = cols[:pinfl_index + 1] + REPORT_COLS + cols[pinfl_index + 1:]
    export_df  = export_df[[c for c in final_cols if c in export_df.columns]]
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Natijalar",       index=False)
        summary.to_excel(  writer, sheet_name="Maktab xulosasi", index=False)
        style_excel(writer)
    out.seek(0)
    return out

# ──────────────────────────────────────────
# PREMIUM CSS
# ──────────────────────────────────────────
PREMIUM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap');

:root {
    --bg: #07111f;
    --panel: rgba(255,255,255,0.075);
    --panel-strong: rgba(255,255,255,0.11);
    --line: rgba(148,163,184,0.20);
    --text: #eef6ff;
    --muted: #9fb0c7;
    --brand: #22d3ee;
    --brand2: #a78bfa;
    --ok: #34d399;
    --warn: #fbbf24;
    --danger: #fb7185;
}

*, *::before, *::after { box-sizing: border-box; }

html, body, [data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 8% 0%, rgba(34,211,238,0.16), transparent 34%),
        radial-gradient(circle at 92% 12%, rgba(167,139,250,0.18), transparent 34%),
        linear-gradient(135deg, #050816 0%, #07111f 48%, #0b1220 100%) !important;
    color: var(--text) !important;
    font-family: 'Inter', sans-serif !important;
}

[data-testid="stAppViewContainer"]::before {
    content: '';
    position: fixed;
    inset: 0;
    pointer-events: none;
    z-index: 0;
    background-image:
        linear-gradient(rgba(255,255,255,0.030) 1px, transparent 1px),
        linear-gradient(90deg, rgba(255,255,255,0.030) 1px, transparent 1px);
    background-size: 42px 42px;
    mask-image: linear-gradient(to bottom, black, transparent 82%);
}

[data-testid="stHeader"] { background: transparent !important; }
[data-testid="stToolbar"] { right: 1rem !important; }
[data-testid="stVerticalBlock"] { position: relative; z-index: 1; }

.block-container {
    max-width: 1180px !important;
    padding: 2rem 1.4rem 4rem !important;
}

/* Header */
.app-hero {
    position: relative;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.14);
    background:
        linear-gradient(135deg, rgba(34,211,238,0.16), rgba(167,139,250,0.13)),
        rgba(255,255,255,0.055);
    box-shadow: 0 24px 80px rgba(0,0,0,0.34);
    border-radius: 30px;
    padding: 34px 36px;
    margin: 0 0 1.5rem;
    backdrop-filter: blur(18px);
}
.app-hero::after {
    content: '';
    position: absolute;
    width: 360px;
    height: 360px;
    right: -120px;
    top: -150px;
    border-radius: 999px;
    background: radial-gradient(circle, rgba(34,211,238,0.30), transparent 65%);
}
.hero-top {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 1rem;
    margin-bottom: 28px;
}
.brand-chip, .safe-chip {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    border-radius: 999px;
    padding: 9px 14px;
    font: 700 12px/1 'JetBrains Mono', monospace;
    letter-spacing: .35px;
    color: #cffafe;
    border: 1px solid rgba(34,211,238,0.30);
    background: rgba(8,47,73,0.42);
}
.safe-chip {
    color: #d1fae5;
    border-color: rgba(52,211,153,0.30);
    background: rgba(6,78,59,0.32);
}
.app-title {
    position: relative;
    z-index: 1;
    max-width: 760px;
    margin: 0;
    font-size: clamp(2.15rem, 5vw, 4.5rem);
    line-height: .96;
    letter-spacing: -0.055em;
    font-weight: 900;
    color: #ffffff;
}
.app-title span {
    background: linear-gradient(135deg, #67e8f9 0%, #c4b5fd 52%, #f0abfc 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.app-subtitle {
    position: relative;
    z-index: 1;
    max-width: 780px;
    margin-top: 18px;
    color: #b7c7da;
    font-size: 1.03rem;
    line-height: 1.65;
}
.hero-grid {
    position: relative;
    z-index: 1;
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 12px;
    margin-top: 28px;
}
.hero-card {
    border: 1px solid rgba(255,255,255,0.13);
    border-radius: 20px;
    background: rgba(2,6,23,0.34);
    padding: 16px;
}
.hero-card b { display:block; color:#ffffff; font-size: 1rem; margin-bottom: 6px; }
.hero-card small { color: var(--muted); line-height: 1.45; }

/* Cards / sections */
.step-card {
    position: relative;
    border: 1px solid rgba(255,255,255,0.12);
    background: rgba(255,255,255,0.065);
    border-radius: 24px;
    padding: 20px 22px;
    margin: 1.25rem 0 1rem;
    box-shadow: 0 18px 54px rgba(0,0,0,0.22);
    backdrop-filter: blur(16px);
}
.step-card::before {
    content: '';
    position: absolute;
    inset: 0;
    border-radius: inherit;
    padding: 1px;
    background: linear-gradient(135deg, rgba(34,211,238,0.38), rgba(167,139,250,0.08), rgba(52,211,153,0.20));
    -webkit-mask: linear-gradient(#000 0 0) content-box, linear-gradient(#000 0 0);
    -webkit-mask-composite: xor;
    mask-composite: exclude;
    pointer-events: none;
}
.step-header { display: flex; align-items: center; gap: 14px; }
.step-num {
    width: 44px; height: 44px;
    border-radius: 16px;
    background: linear-gradient(135deg, #22d3ee, #8b5cf6);
    display: flex; align-items: center; justify-content: center;
    color: white; font-weight: 900;
    box-shadow: 0 10px 30px rgba(34,211,238,0.20);
}
.step-title { color: #f8fafc; font-size: 1.14rem; font-weight: 850; letter-spacing: -0.015em; }
.step-desc { color: var(--muted); margin-top: 4px; font-size: .92rem; }

/* Native Streamlit elements */
[data-testid="stExpander"] {
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 20px !important;
    background: rgba(255,255,255,0.055) !important;
    overflow: hidden;
    box-shadow: 0 14px 44px rgba(0,0,0,0.18);
}
[data-testid="stExpander"] summary { font-weight: 750 !important; color: #e0f2fe !important; }

.stTextArea textarea, .stTextInput input {
    background: rgba(2,6,23,0.72) !important;
    color: #e2e8f0 !important;
    border: 1px solid rgba(148,163,184,0.25) !important;
    border-radius: 18px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 12.5px !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.05) !important;
}
.stTextArea textarea:focus, .stTextInput input:focus {
    border-color: rgba(34,211,238,0.68) !important;
    box-shadow: 0 0 0 4px rgba(34,211,238,0.10) !important;
}

[data-testid="stFileUploader"] {
    border: 1.5px dashed rgba(34,211,238,0.40) !important;
    background: rgba(34,211,238,0.055) !important;
    border-radius: 22px !important;
    padding: 1rem !important;
}
[data-testid="stFileUploader"] section {
    background: rgba(2,6,23,0.34) !important;
    border-radius: 16px !important;
}

.stButton > button, .stDownloadButton > button {
    width: 100%;
    border: 0 !important;
    border-radius: 18px !important;
    color: #ffffff !important;
    background: linear-gradient(135deg, #06b6d4 0%, #7c3aed 54%, #db2777 100%) !important;
    box-shadow: 0 14px 34px rgba(124,58,237,0.30) !important;
    padding: .86rem 1.35rem !important;
    font-weight: 850 !important;
    letter-spacing: -.01em !important;
    transition: transform .18s ease, box-shadow .18s ease, filter .18s ease !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-2px) !important;
    filter: brightness(1.08) !important;
    box-shadow: 0 20px 44px rgba(6,182,212,0.28) !important;
}
.stDownloadButton > button { background: linear-gradient(135deg, #10b981, #06b6d4) !important; }

[data-testid="metric-container"] {
    border: 1px solid rgba(255,255,255,0.13) !important;
    background: rgba(255,255,255,0.065) !important;
    border-radius: 22px !important;
    padding: 1.15rem 1.25rem !important;
    box-shadow: 0 14px 40px rgba(0,0,0,0.16);
}
[data-testid="stMetricLabel"] { color: #a9b8cc !important; font-weight: 700 !important; }
[data-testid="stMetricValue"] { color: #ffffff !important; font-size: 2.1rem !important; font-weight: 900 !important; letter-spacing: -0.04em; }

.status-bar {
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
    color: #cffafe;
    background: rgba(8,47,73,0.46);
    border: 1px solid rgba(34,211,238,0.22);
    border-radius: 16px;
    padding: 12px 16px;
    margin: 10px 0;
}
[data-testid="stProgress"] > div { background: rgba(15,23,42,0.85) !important; border-radius: 999px !important; height: 10px !important; }
[data-testid="stProgress"] > div > div { background: linear-gradient(90deg, #22d3ee, #a78bfa, #34d399) !important; border-radius: 999px !important; }

[data-testid="stSuccess"], [data-testid="stInfo"], [data-testid="stWarning"], [data-testid="stError"] {
    border-radius: 16px !important;
    border: 1px solid rgba(255,255,255,0.13) !important;
    backdrop-filter: blur(12px);
}
[data-testid="stSuccess"] { background: rgba(16,185,129,0.12) !important; color: #bbf7d0 !important; }
[data-testid="stInfo"] { background: rgba(14,165,233,0.12) !important; color: #bae6fd !important; }
[data-testid="stWarning"] { background: rgba(245,158,11,0.13) !important; color: #fde68a !important; }
[data-testid="stError"] { background: rgba(244,63,94,0.13) !important; color: #fecdd3 !important; }

[data-testid="stDataFrame"] {
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 22px !important;
    overflow: hidden !important;
    box-shadow: 0 18px 54px rgba(0,0,0,0.20);
}
[data-testid="stMultiSelect"] div { border-radius: 14px !important; }
hr { border-color: rgba(148,163,184,0.16) !important; margin: 1.8rem 0 !important; }
h1, h2, h3, .stMarkdown { color: var(--text); }

.app-footer {
    text-align: center;
    margin-top: 3rem;
    padding: 28px 0 12px;
    color: #7c8ba1;
}
.footer-badge {
    display: inline-flex;
    align-items: center;
    gap: 10px;
    border: 1px solid rgba(255,255,255,0.12);
    background: rgba(255,255,255,0.055);
    border-radius: 999px;
    padding: 10px 18px;
    font: 600 12px/1 'JetBrains Mono', monospace;
}
.footer-badge span { color: #67e8f9; }

@media (max-width: 760px) {
    .block-container { padding: 1rem .7rem 3rem !important; }
    .app-hero { padding: 24px 20px; border-radius: 24px; }
    .hero-top { align-items: flex-start; flex-direction: column; }
    .hero-grid { grid-template-columns: 1fr; }
}
</style>
"""

# ──────────────────────────────────────────
# STREAMLIT APP
# ──────────────────────────────────────────
st.set_page_config(
    page_title="AI Leaders PINFL Checker",
    page_icon="🎓",
    layout="wide",
)

st.markdown(PREMIUM_CSS, unsafe_allow_html=True)

# ── HEADER ──
st.markdown("""
<div class="app-hero">
    <div class="hero-top">
        <div class="brand-chip">⚡ AI Leaders · PINFL Checker</div>
        <div class="safe-chip">🛡️ Excel + Cookie based check</div>
    </div>
    <h1 class="app-title">Sertifikatlarni <span>tez va aniq</span> tekshirish</h1>
    <div class="app-subtitle">aileaders.uz bazasi bo‘yicha PINFL ro‘yxatlarini avtomatik tekshiradi, maktab kesimida xulosa chiqaradi va tayyor Excel hisobot beradi.</div>
    <div class="hero-grid">
        <div class="hero-card"><b>01 · cURL</b><small>Saytdan autentifikatsiya ma’lumoti olinadi.</small></div>
        <div class="hero-card"><b>02 · Excel</b><small>PINFL ustunlari avtomatik aniqlanadi.</small></div>
        <div class="hero-card"><b>03 · Hisobot</b><small>Natija va maktab xulosasi eksport qilinadi.</small></div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── 1-QADAM: cURL ──
st.markdown("""
<div class="step-card">
    <div class="step-header">
        <div class="step-num">1</div>
        <div><div class="step-title">cURL — Autentifikatsiya</div><div class="step-desc">Chrome Network’dan olingan cURL matnini joylashtiring.</div></div>
    </div>
</div>
""", unsafe_allow_html=True)

with st.expander("📋 cURL qanday olish kerak?", expanded=True):
    st.markdown("""
**Chrome da bir marta bajaring:**

1. `https://aileaders.uz/auth/login/check` sahifasini oching
2. Istalgan PINFL kiriting → **Tekshirish** bosing
3. **F12** → **Network** tab
4. `certificates?pinfl=...` qatoriga **o'ng klik**
5. **Copy → Copy as cURL (bash)** tanlang
6. Quyidagi maydonga **Ctrl+V** bilan joylashtiring
    """)

curl_input = st.text_area(
    "cURL matni:",
    placeholder="curl 'https://aileaders.uz/api/v1/check/certificates?pinfl=...' \\\n  -b 'HWWAFSESID=...; HWWAFSESTIME=...' \\\n  -H 'user-agent: Mozilla/5.0 ...'",
    height=110,
    label_visibility="collapsed",
)

parsed  = {}
curl_ok = False

if curl_input.strip():
    parsed = parse_curl(curl_input)
    if parsed["cookie"] and "HWWAFSESID" in parsed["cookie"]:
        st.success("✅ Cookie muvaffaqiyatli aniqlandi — tayyor!")
        curl_ok = True
    else:
        st.error("❌ Cookie topilmadi. cURL to'liq ko'chirilganmi?")

st.divider()

# ── 2-QADAM: EXCEL ──
st.markdown("""
<div class="step-card">
    <div class="step-header">
        <div class="step-num">2</div>
        <div><div class="step-title">Excel — O'quvchilar ro'yxati</div><div class="step-desc">.xlsx fayl yuklang, tizim PINFL ustunini o‘zi topadi.</div></div>
    </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Excel fayl (.xlsx)",
    type=["xlsx"],
    label_visibility="collapsed",
)

if uploaded:
    with st.spinner("Fayl tahlil qilinmoqda..."):
        df_all = read_excel(uploaded)

    if df_all.empty:
        st.error("❌ Excel dan PINFL ustuni topilmadi!")
        st.stop()

    col_a, col_b = st.columns(2)
    col_a.success(f"✅ **{len(df_all):,}** ta o'quvchi yuklandi")
    col_b.info(f"🏫 **{df_all['Maktab'].nunique()}** ta maktab aniqlandi")

    with st.expander("🏫 Maktab filtri (ixtiyoriy)"):
        maktablar = sorted(df_all["Maktab"].unique().tolist())
        tanlangan = st.multiselect(
            "Faqat quyidagi maktablarni tekshirish (bo'sh = hammasi)",
            maktablar,
        )
        if tanlandan := tanlangan:
            df_all = df_all[df_all["Maktab"].isin(tanlandan)].copy()
            st.info(f"Tanlangan: **{len(df_all)}** ta o'quvchi")

    daqiqa = round(len(df_all) * DELAY_SEC / 60, 1)
    soat   = round(daqiqa / 60, 1)
    st.info(
        f"⏱️ Taxminiy vaqt: **{daqiqa} daqiqa** (~{soat} soat) · "
        f"{len(df_all):,} ta × {DELAY_SEC}s · Rate limit: 20 req/min"
    )

    st.divider()

    # ── 3-QADAM: TEKSHIRISH ──
    st.markdown("""
    <div class="step-card">
        <div class="step-header">
            <div class="step-num">3</div>
            <div><div class="step-title">Tekshirishni boshlash</div><div class="step-desc">Jarayon progress, status va yakuniy Excel eksport bilan ishlaydi.</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if not curl_ok:
        st.warning("⚠️ Avval 1-qadamda cURL ni joylashtiring!")

    if curl_ok and st.button("🚀 Tekshirishni boshlash", type="primary"):

        session = requests.Session()
        session.headers.update({
            "User-Agent": parsed.get("user_agent") or (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36"
            ),
            "Accept":   "*/*",
            "Referer":  "https://aileaders.uz/auth/login/check",
            "Cookie":   parsed["cookie"],
        })

        progress_bar = st.progress(0)
        status_el    = st.empty()
        results      = []
        total        = len(df_all)
        cookie_dead  = False

        name_col = next(
            (c for c in df_all.columns if "наименование" in str(c).lower()), None
        )

        for i, (_, row) in enumerate(df_all.iterrows(), 1):
            pinfl  = str(row["PINFL"])
            maktab = str(row["Maktab"])

            status_el.markdown(
                f'<div class="status-bar">🔄 {i}/{total} — <b>{pinfl}</b> · {maktab}</div>',
                unsafe_allow_html=True,
            )
            progress_bar.progress(i / total)

            # Rate limit bo'lsa qayta urinish
            res = check_pinfl_safe(pinfl, session, status_el, i, total)

            if res["holat"] == "🔐 Cookie eskirgan":
                st.error("🔐 Cookie eskirdi! Yangi cURL olib, 1-qadamdan qayta boshlang.")
                cookie_dead = True
                break

            fio = str(row.get(name_col, "") or "") if name_col else ""

            results.append({
                "_RID_":               row["_RID_"],
                "Maktab":              maktab,
                "F.I.Sh.":             fio,
                "PINFL":               pinfl,
                "Tekshiruv holati":    res["holat"],
                "Saytdagi F.I.Sh.":   res["ism"],
                "Email":               res["email"],
                "Kurslar soni":        res["kurslar"],
                "Yakunlangan kurslar": res["yakunlangan"],
                "Kurslar tafsiloti":   res["kurs_tafsiloti"],
                "Izoh":                res["xato"],
            })
            time.sleep(DELAY_SEC)

        if not cookie_dead:
            status_el.markdown(
                f'<div class="status-bar">✅ Yakunlandi — {len(results)}/{total} ta tekshirildi</div>',
                unsafe_allow_html=True,
            )
        progress_bar.progress(1.0)

        if not results:
            st.stop()

        result_df = pd.DataFrame(results)

        # ── STATISTIKA ──
        st.subheader("📊 Natijalar")
        c1, c2, c3, c4 = st.columns(4)
        n_total    = len(result_df)
        n_sert     = (result_df["Tekshiruv holati"] == "✅ Sertifikat olgan").sum()
        n_kurs     = result_df["Tekshiruv holati"].str.contains("Kurs bor|kurs topilmadi", case=False, na=False).sum()
        n_yoq      = (result_df["Tekshiruv holati"] == "❌ Ro'yxatdan o'tmagan").sum()

        c1.metric("Jami tekshirildi",    f"{n_total:,}")
        c2.metric("✅ Sertifikat olgan", f"{n_sert:,}")
        c3.metric("⚠️ Sertifikatsiz",   f"{n_kurs:,}")
        c4.metric("❌ Ro'yxatda yo'q",  f"{n_yoq:,}")

        st.dataframe(
            result_df.drop(columns=["_RID_"]),
            use_container_width=True,
            height=400,
        )

        # ── MAKTAB XULOSASI ──
        st.subheader("🏫 Maktab bo'yicha hisobot")
        summary = (
            result_df.groupby("Maktab")
            .agg(
                Jami=("PINFL", "count"),
                Sertifikat_olgan=("Tekshiruv holati", lambda x: (x == "✅ Sertifikat olgan").sum()),
                Sertifikatsiz=("Tekshiruv holati", lambda x: x.str.contains("Kurs bor|kurs topilmadi", case=False, na=False).sum()),
                Royxatdan_otmagan=("Tekshiruv holati", lambda x: (x == "❌ Ro'yxatdan o'tmagan").sum()),
            )
            .assign(Foiz=lambda d: (d["Sertifikat_olgan"] / d["Jami"] * 100).round(1))
            .sort_values("Foiz", ascending=False)
            .reset_index()
        )
        st.dataframe(summary, use_container_width=True)

        # ── YUKLAB OLISH ──
        with st.spinner("Excel hisobot tayyorlanmoqda..."):
            out = build_report_excel(df_all, result_df, summary)

        st.download_button(
            "📥 To'liq hisobotni yuklab olish (Excel)",
            data=out,
            file_name="aileaders_hisobot.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ── FOOTER ──
st.markdown("""
<div class="app-footer">
    <div class="footer-badge">
        Developed by <span>Azamat Madrimov</span> &nbsp;·&nbsp; InfoSchoolUz Khorezm &nbsp;·&nbsp; 2026
    </div>
</div>
""", unsafe_allow_html=True)
